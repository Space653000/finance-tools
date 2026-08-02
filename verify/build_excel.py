# -*- coding: utf-8 -*-
"""產生 Excel 交叉驗算表。
   儲存格填的是 Excel 原生財務函式（PMT / FV / IPMT / PPMT / RATE），
   由 Excel 自己算，是完全獨立於本專案程式的第三方來源。"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = os.path.dirname(os.path.abspath(__file__))
js = json.load(open(os.path.join(D, "js_results.json"), encoding="utf-8"))["cases"]

wb = Workbook()
H  = Font(bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="0A7C74")
T  = Font(bold=True, size=13)
MONO = Font(name="Consolas", size=10)
thin = Side(style="thin", color="D0D0D0")
BD = Border(left=thin, right=thin, top=thin, bottom=thin)

def head(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row, i, c); cell.font = H; cell.fill = HF
        cell.alignment = Alignment(horizontal="center"); cell.border = BD
def widths(ws, w):
    for i, x in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = x

# ================= 工作表 1：攤還 =================
ws = wb.active; ws.title = "攤還驗算"
ws["A1"] = "借款攤還　Excel 原生函式 vs 網頁引擎"; ws["A1"].font = T
ws["A2"] = "C 欄為 Excel 自行計算（PMT 函式），E 欄為網頁引擎輸出，F 欄為差異。差異應為 0。"
widths(ws, [30, 14, 18, 14, 18, 14, 12])
head(ws, 4, ["情境", "本金", "Excel PMT", "年利率", "網頁引擎", "差異", "判定"])

cases = [
 ("房貸 500 萬・2.35%・30 年", 5000000, 2.35, 30, "pmt_mortgage_5M_2p35_30y"),
 ("信貸 100 萬・5.5%・7 年",   1000000, 5.5,  7,  "pmt_credit_1M_5p5_7y"),
 ("零利率 120 萬・10 年",      1200000, 0,    10, "edge_zero_rate_pmt"),
 ("寬限 3 年後 500 萬・27 年",  5000000, 2.35, 27, "pmt_after_grace"),
]
r = 5
for name, P, rate, yrs, key in cases:
    ws.cell(r,1,name).border = BD
    ws.cell(r,2,P).border = BD; ws.cell(r,2).number_format = "#,##0"
    ws.cell(r,4,rate/100).border = BD; ws.cell(r,4).number_format = "0.00%"
    ws.cell(r,3, f"=IF(D{r}=0, B{r}/{yrs*12}, -PMT(D{r}/12,{yrs*12},B{r}))").border = BD
    ws.cell(r,3).number_format = "#,##0.00"
    ws.cell(r,5, round(js[key],6)).border = BD; ws.cell(r,5).number_format = "#,##0.00"
    ws.cell(r,6, f"=C{r}-E{r}").border = BD; ws.cell(r,6).number_format = "0.000000"
    ws.cell(r,7, f'=IF(ABS(F{r})<0.01,"✅ 相符","❌ 不符")').border = BD
    r += 1

r += 2
ws.cell(r,1,"攤還首期拆分（Excel IPMT / PPMT）").font = T; r += 1
head(ws, r, ["項目", "Excel 計算", "網頁引擎", "差異", "判定"]); r += 1
for label, formula, key in [
  ("第 1 期利息", "=-IPMT(0.0235/12,1,360,5000000)", "amort_m1_interest"),
  ("第 1 期本金", "=-PPMT(0.0235/12,1,360,5000000)", "amort_m1_principal"),
  ("第 12 期後餘額", "=-FV(0.0235/12,12,PMT(0.0235/12,360,5000000),5000000)", "amort_balance_after_12m"),
]:
    ws.cell(r,1,label).border = BD
    ws.cell(r,2,formula).border = BD; ws.cell(r,2).number_format = "#,##0.00"
    ws.cell(r,3,round(js[key],6)).border = BD; ws.cell(r,3).number_format = "#,##0.00"
    ws.cell(r,4,f"=B{r}-C{r}").border = BD; ws.cell(r,4).number_format = "0.000000"
    ws.cell(r,5,f'=IF(ABS(D{r})<0.01,"✅ 相符","❌ 不符")').border = BD
    r += 1

# ================= 工作表 2：複利 =================
ws2 = wb.create_sheet("複利驗算")
ws2["A1"] = "複利終值　Excel FV 函式 vs 網頁引擎"; ws2["A1"].font = T
ws2["A2"] = "B 欄由 Excel 的 FV 函式自行計算。年複利以年為期，月複利以月為期。"
widths(ws2, [34, 20, 20, 16, 12])
head(ws2, 4, ["情境", "Excel FV", "網頁引擎", "差異", "判定"])
comp = [
 ("單筆 100 萬・6%・10 年・年複利", "=-FV(0.06,10,0,1000000)",                         "lump_1M_6pct_10y_annual"),
 ("單筆 100 萬・6%・10 年・月複利", "=-FV((1+0.06/12)^12-1,10,0,1000000)",             "lump_1M_6pct_10y_monthly"),
 ("單筆 500 萬・8%・20 年・年複利", "=-FV(0.08,20,0,5000000)",                         "lump_5M_8pct_20y_annual"),
 ("月投 1 萬・6%・10 年",           "=-FV((1+0.06/12)^(12/12)-1,120,10000,0)",         "dca_10k_6pct_10y"),
 ("月投 3 萬・5%・30 年",           "=-FV((1+0.05/12)^(12/12)-1,360,30000,0)",         "dca_30k_5pct_30y"),
 ("混合 100 萬＋月投 1 萬・6%・10 年","=-FV((1+0.06/12)^(12/12)-1,120,10000,1000000)",  "mix_1M_10k_6pct_10y"),
]
r = 5
for name, formula, key in comp:
    ws2.cell(r,1,name).border = BD
    ws2.cell(r,2,formula).border = BD; ws2.cell(r,2).number_format = "#,##0.00"
    ws2.cell(r,3,round(js[key],6)).border = BD; ws2.cell(r,3).number_format = "#,##0.00"
    ws2.cell(r,4,f"=B{r}-C{r}").border = BD; ws2.cell(r,4).number_format = "0.00"
    ws2.cell(r,5,f'=IF(ABS(D{r})<1,"✅ 相符","❌ 不符")').border = BD
    r += 1

# ================= 工作表 3：擔保與稅 =================
ws3 = wb.create_sheet("擔保與稅")
ws3["A1"] = "擔保維持率與稅務　公式自行驗算"; ws3["A1"].font = T
widths(ws3, [34, 22, 20, 16, 12])
ws3["A3"] = "擔保品市值"; ws3["B3"] = 5000000; ws3["B3"].number_format = "#,##0"
ws3["A4"] = "擔保借款";   ws3["B4"] = 2000000; ws3["B4"].number_format = "#,##0"
ws3["A5"] = "追繳維持率"; ws3["B5"] = 1.40;    ws3["B5"].number_format = "0%"
ws3["A6"] = "處分維持率"; ws3["B6"] = 1.30;    ws3["B6"].number_format = "0%"
ws3["A7"] = "可動用現金"; ws3["B7"] = 600000;  ws3["B7"].number_format = "#,##0"
ws3["A8"] = "每月付息";   ws3["B8"] = "=B4*0.032/12"; ws3["B8"].number_format = "#,##0.00"
head(ws3, 10, ["項目", "Excel 公式計算", "網頁引擎", "差異", "判定"])
mg = [
 ("擔保維持率 %",        "=B3/B4*100",          "margin_ratio"),
 ("距離追繳 %",          "=(1-B5*B4/B3)*100",   "margin_to_call"),
 ("距離處分 %",          "=(1-B6*B4/B3)*100",   "margin_to_sell"),
 ("現金可支應月數",      "=B7/B8",              "margin_months"),
]
r = 11
for name, formula, key in mg:
    ws3.cell(r,1,name).border = BD
    ws3.cell(r,2,formula).border = BD; ws3.cell(r,2).number_format = "#,##0.0000"
    ws3.cell(r,3,round(js[key],6)).border = BD; ws3.cell(r,3).number_format = "#,##0.0000"
    ws3.cell(r,4,f"=B{r}-C{r}").border = BD; ws3.cell(r,4).number_format = "0.000000"
    ws3.cell(r,5,f'=IF(ABS(D{r})<0.0001,"✅ 相符","❌ 不符")').border = BD
    r += 1

r += 2
ws3.cell(r,1,"股利有效稅率").font = T; r += 1
head(ws3, r, ["課稅方式", "Excel 公式計算", "網頁引擎", "差異", "判定"]); r += 1
for name, formula, key in [
 ("合併計稅・稅率 12%", "=MAX(0,0.12-0.085)", "tax_combined_12pct"),
 ("合併計稅・稅率 40%", "=MAX(0,0.40-0.085)", "tax_combined_40pct"),
 ("分開計稅 28%",       "=0.28",              "tax_separate"),
]:
    ws3.cell(r,1,name).border = BD
    ws3.cell(r,2,formula).border = BD; ws3.cell(r,2).number_format = "0.0000"
    ws3.cell(r,3,round(js[key],6)).border = BD; ws3.cell(r,3).number_format = "0.0000"
    ws3.cell(r,4,f"=B{r}-C{r}").border = BD; ws3.cell(r,4).number_format = "0.000000"
    ws3.cell(r,5,f'=IF(ABS(D{r})<0.000001,"✅ 相符","❌ 不符")').border = BD
    r += 1

# ================= 工作表 4：真實市場條件 =================
ws4 = wb.create_sheet("真實條件試算")
ws4["A1"] = "以真實市場常見條件試算　供人工比對合理性"; ws4["A1"].font = T
ws4["A2"] = "利率為市場常見區間，非任何機構報價。實際條件請以各金融機構公告與合約為準。"
widths(ws4, [26, 14, 12, 10, 18, 20, 26])
head(ws4, 4, ["借款類型", "金額", "年利率", "年期", "Excel 月付金", "20 年總利息", "備註"])
real = [
 ("不限用途款項借貸", 2000000, 3.2,  1,  "按期付息、到期一次還本"),
 ("證券融資",         2000000, 6.4,  1,  "整戶維持率低於 130% 追繳"),
 ("房屋貸款",         5000000, 2.35, 30, "機動利率，升息月付金增加"),
 ("理財型房屋貸款",   3000000, 3.0,  20, "隨借隨還、按日計息"),
 ("個人信用貸款",     1000000, 5.5,  7,  "計入 DBR 22 倍上限"),
 ("汽車貸款",          800000, 5.5,  5,  "車輛折舊快於本金攤還"),
]
r = 5
for name, P, rate, yrs, note in real:
    ws4.cell(r,1,name).border = BD
    ws4.cell(r,2,P).border = BD; ws4.cell(r,2).number_format = "#,##0"
    ws4.cell(r,3,rate/100).border = BD; ws4.cell(r,3).number_format = "0.00%"
    ws4.cell(r,4,yrs).border = BD
    if yrs <= 1:
        ws4.cell(r,5, f"=B{r}*C{r}/12").border = BD
        ws4.cell(r,6, f"=B{r}*C{r}*20").border = BD
    else:
        ws4.cell(r,5, f"=-PMT(C{r}/12,D{r}*12,B{r})").border = BD
        ws4.cell(r,6, f"=E{r}*MIN(D{r},20)*12-B{r}*MIN(D{r},20)/D{r}").border = BD
    ws4.cell(r,5).number_format = "#,##0"; ws4.cell(r,6).number_format = "#,##0"
    ws4.cell(r,7,note).border = BD
    r += 1
r += 1
ws4.cell(r,1,"合計月付金").font = T
ws4.cell(r,5,f"=SUM(E5:E{r-2})").font = T; ws4.cell(r,5).number_format = "#,##0"
r += 1
ws4.cell(r,1,"月收入 12 萬時的月付金佔比")
ws4.cell(r,5,f"=E{r-1}/120000"); ws4.cell(r,5).number_format = "0.0%"
r += 1
ws4.cell(r,1,"銀行實務警戒值")
ws4.cell(r,5,0.40); ws4.cell(r,5).number_format = "0%"
r += 1
ws4.cell(r,1,"判定")
ws4.cell(r,5,f'=IF(E{r-2}<=E{r-1},"✅ 未超出警戒","❌ 已超出警戒")')

# ================= 工作表 5：第三方 App 對照 =================
ws5 = wb.create_sheet("第三方對照")
ws5["A1"] = "與市售房貸試算 App 交叉比對"; ws5["A1"].font = T
ws5["A2"] = "條件：500 萬・年利率 2.8%・20 年・寬限期 3 年。B 欄由 Excel 自算，C 欄為對照 App 顯示值。"
widths(ws5, [30, 20, 20, 20, 16, 12])
head(ws5, 4, ["項目", "Excel 計算", "對照 App", "本工具", "差異", "判定"])
ref = [
 ("無寬限　月付金",        "=-PMT(0.028/12,240,5000000)", 27232,    "ref_500w_2p8_20y_normal"),
 ("寬限期內　月付金",      "=5000000*0.028/12",           11666.67, "ref_500w_2p8_20y_grace"),
 ("第 37 月起　月付金",    "=-PMT(0.028/12,204,5000000)", 30832,    "ref_500w_2p8_20y_after"),
]
r = 5
for name, formula, app_val, key in ref:
    ws5.cell(r,1,name).border = BD
    ws5.cell(r,2,formula).border = BD; ws5.cell(r,2).number_format = "#,##0.00"
    ws5.cell(r,3,app_val).border = BD; ws5.cell(r,3).number_format = "#,##0.00"
    ws5.cell(r,4,round(js[key],2)).border = BD; ws5.cell(r,4).number_format = "#,##0.00"
    ws5.cell(r,5,f"=B{r}-D{r}").border = BD; ws5.cell(r,5).number_format = "0.0000"
    ws5.cell(r,6,f'=IF(ABS(E{r})<0.01,"✅ 相符","❌ 不符")').border = BD
    r += 1
r += 1
ws5.cell(r,1,"寬限期滿後月付金跳增").font = T
ws5.cell(r,2,"=B7-B6"); ws5.cell(r,2).number_format = "#,##0.00"
ws5.cell(r,4,"=D7-D6"); ws5.cell(r,4).number_format = "#,##0.00"
r += 1
ws5.cell(r,1,"說明")
ws5.cell(r,2,"寬限期內僅繳利息，期滿後以剩餘 17 年重新計算本息平均攤還")

wb.save(os.path.join(D, "verification_workbook.xlsx"))
print("✓ 已產生 驗算表.xlsx（4 個工作表，公式由 Excel 自行計算）")
